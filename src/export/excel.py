"""Excel export — generate the linking plan spreadsheet."""

from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Brand colors
GREEN = "005947"
LIGHT_GREEN = "E8F5F1"
RED = "D63637"
LIGHT_RED = "FDE8E8"
YELLOW = "B8860B"
LIGHT_YELLOW = "FDF3DB"
GRAY = "6C757D"
LIGHT_GRAY = "F8F9FA"
WHITE = "FFFFFF"

HEADER_FONT = Font(name="Roboto", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
SUBHEADER_FONT = Font(name="Roboto", bold=True, color=GREEN, size=10)
BODY_FONT = Font(name="Roboto", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E5E9"),
)

STATUS_FILLS = {
    "live": PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"),
    "to add": PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"),
    "to remove": PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid"),
}

STATUS_FONTS = {
    "live": Font(name="Roboto", size=10, color=GRAY),
    "to add": Font(name="Roboto", size=10, color=GREEN, bold=True),
    "to remove": Font(name="Roboto", size=10, color=RED, bold=True),
}


def generate_linking_plan(
    cleaned_df: pd.DataFrame,
    recommendations: list[dict],
    audit_results: dict,
    priority_health_df: pd.DataFrame | None = None,
    cocoon_health_df: pd.DataFrame | None = None,
    token_usage: dict | None = None,
) -> bytes:
    """Generate the Excel linking plan spreadsheet.

    Contains two sheets:
    1. Linking Plan — all links with status (live / to add / to remove)
    2. Summary — audit stats, priority URL health, cocoon health

    Returns:
        Excel file as bytes (ready for st.download_button).
    """
    wb = Workbook()

    # ---- Sheet 1: Linking Plan ----
    ws_plan = wb.active
    ws_plan.title = "Linking Plan"

    # Header row
    headers = ["Source URL", "Anchor", "Target URL", "Status", "Priority", "Reason"]
    for col, header in enumerate(headers, 1):
        cell = ws_plan.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Set column widths
    ws_plan.column_dimensions["A"].width = 55  # Source URL
    ws_plan.column_dimensions["B"].width = 35  # Anchor
    ws_plan.column_dimensions["C"].width = 55  # Target URL
    ws_plan.column_dimensions["D"].width = 12  # Status
    ws_plan.column_dimensions["E"].width = 10  # Priority
    ws_plan.column_dimensions["F"].width = 50  # Reason

    row_num = 2

    # AI recommendations ("to add")
    for rec in recommendations:
        source = rec.get("source_url", "")
        anchor = rec.get("suggested_anchor", "")
        target = rec.get("target_url", "")
        priority = rec.get("priority", "")
        reason = rec.get("reason", "")

        ws_plan.cell(row=row_num, column=1, value=source).font = BODY_FONT
        ws_plan.cell(row=row_num, column=2, value=anchor).font = BODY_FONT
        ws_plan.cell(row=row_num, column=3, value=target).font = BODY_FONT

        status_cell = ws_plan.cell(row=row_num, column=4, value="to add")
        status_cell.font = STATUS_FONTS["to add"]
        status_cell.fill = STATUS_FILLS["to add"]

        ws_plan.cell(row=row_num, column=5, value=priority).font = BODY_FONT
        ws_plan.cell(row=row_num, column=6, value=reason).font = BODY_FONT

        for col in range(1, 7):
            ws_plan.cell(row=row_num, column=col).border = THIN_BORDER
            ws_plan.cell(row=row_num, column=col).alignment = Alignment(vertical="center", wrap_text=True)

        row_num += 1

    # Existing links ("live") — from cleaned data, only for priority URL targets
    if priority_health_df is not None:
        priority_urls = set(priority_health_df["URL"].tolist())
        live_links = cleaned_df[cleaned_df["Destination"].isin(priority_urls)]

        for _, link_row in live_links.iterrows():
            source = link_row["Source"]
            anchor = link_row.get("Anchor", "")
            target = link_row["Destination"]

            ws_plan.cell(row=row_num, column=1, value=source).font = BODY_FONT
            ws_plan.cell(row=row_num, column=2, value=anchor).font = BODY_FONT
            ws_plan.cell(row=row_num, column=3, value=target).font = BODY_FONT

            status_cell = ws_plan.cell(row=row_num, column=4, value="live")
            status_cell.font = STATUS_FONTS["live"]
            status_cell.fill = STATUS_FILLS["live"]

            ws_plan.cell(row=row_num, column=5, value="").font = BODY_FONT
            ws_plan.cell(row=row_num, column=6, value="Existing link").font = BODY_FONT

            for col in range(1, 7):
                ws_plan.cell(row=row_num, column=col).border = THIN_BORDER
                ws_plan.cell(row=row_num, column=col).alignment = Alignment(vertical="center", wrap_text=True)

            row_num += 1

    # Freeze header row
    ws_plan.freeze_panes = "A2"
    # Auto-filter
    ws_plan.auto_filter.ref = f"A1:F{max(row_num - 1, 1)}"

    # ---- Sheet 2: Summary ----
    ws_summary = wb.create_sheet("Summary")

    # Title
    cell = ws_summary.cell(row=1, column=1, value="Internal Link Analysis — Summary")
    cell.font = Font(name="Roboto", bold=True, color=GREEN, size=14)
    ws_summary.merge_cells("A1:D1")

    date_cell = ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    date_cell.font = Font(name="Roboto", size=10, color=GRAY)

    # Audit stats
    row = 4
    ws_summary.cell(row=row, column=1, value="Site Overview").font = SUBHEADER_FONT
    row += 1

    stats = [
        ("Total Pages", f"{audit_results['total_pages']:,}"),
        ("Total Internal Links", f"{audit_results['total_links']:,}"),
        ("Orphan Pages", f"{audit_results['orphan_count']:,}"),
        ("True Orphan Pages", f"{audit_results.get('true_orphan_count', 0):,}"),
        ("Avg Inbound Links/Page", str(audit_results["inbound_avg"])),
        ("Median Inbound Links/Page", str(audit_results["inbound_median"])),
        ("Max Inbound Links", str(audit_results["inbound_max"])),
        ("Avg Outbound Links/Page", str(audit_results["outbound_avg"])),
    ]

    for label, value in stats:
        ws_summary.cell(row=row, column=1, value=label).font = BODY_FONT
        ws_summary.cell(row=row, column=2, value=value).font = Font(name="Roboto", bold=True, size=10)
        row += 1

    # Recommendations summary
    row += 1
    ws_summary.cell(row=row, column=1, value="AI Recommendations").font = SUBHEADER_FONT
    row += 1

    n_recs = len(recommendations)
    n_high = sum(1 for r in recommendations if r.get("priority") == "high")
    n_med = sum(1 for r in recommendations if r.get("priority") == "medium")
    n_low = sum(1 for r in recommendations if r.get("priority") == "low")

    rec_stats = [
        ("Total Recommendations", str(n_recs)),
        ("High Priority", str(n_high)),
        ("Medium Priority", str(n_med)),
        ("Low Priority", str(n_low)),
    ]

    for label, value in rec_stats:
        ws_summary.cell(row=row, column=1, value=label).font = BODY_FONT
        ws_summary.cell(row=row, column=2, value=value).font = Font(name="Roboto", bold=True, size=10)
        row += 1

    # Priority URLs health
    if priority_health_df is not None and not priority_health_df.empty:
        row += 1
        ws_summary.cell(row=row, column=1, value="Priority URLs Health").font = SUBHEADER_FONT
        row += 1

        health_headers = ["URL", "Target Keyword", "Inbound Links", "Health"]
        for col, header in enumerate(health_headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1

        for _, prow in priority_health_df.iterrows():
            ws_summary.cell(row=row, column=1, value=prow["URL"]).font = BODY_FONT
            ws_summary.cell(row=row, column=2, value=prow["Target Keyword"]).font = BODY_FONT
            ws_summary.cell(row=row, column=3, value=prow["Inbound Links"]).font = BODY_FONT
            health_val = prow["Health"]
            health_cell = ws_summary.cell(row=row, column=4, value=health_val)
            health_cell.font = BODY_FONT
            if health_val == "critical":
                health_cell.fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
            elif health_val == "warning":
                health_cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
            elif health_val == "good":
                health_cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
            row += 1

    # Cocoon health
    if cocoon_health_df is not None and not cocoon_health_df.empty:
        row += 1
        ws_summary.cell(row=row, column=1, value="Cocoon Health").font = SUBHEADER_FONT
        row += 1

        cocoon_headers = ["Operator", "Pages", "Intra-links", "Completeness %", "Code Page Links", "Health"]
        for col, header in enumerate(cocoon_headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1

        for _, crow in cocoon_health_df.iterrows():
            ws_summary.cell(row=row, column=1, value=crow["Operator"]).font = BODY_FONT
            ws_summary.cell(row=row, column=2, value=crow["Pages"]).font = BODY_FONT
            ws_summary.cell(row=row, column=3, value=f"{crow['Intra-links']} / {crow['Max Possible']}").font = BODY_FONT
            ws_summary.cell(row=row, column=4, value=crow["Completeness"]).font = BODY_FONT
            ws_summary.cell(row=row, column=5, value=crow["Code Page Links"]).font = BODY_FONT
            health_val = crow["Health"]
            health_cell = ws_summary.cell(row=row, column=6, value=health_val)
            health_cell.font = BODY_FONT
            if health_val == "poor":
                health_cell.fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
            elif health_val == "weak":
                health_cell.fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
            elif health_val == "good":
                health_cell.fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
            row += 1

    # Token usage
    if token_usage and token_usage.get("api_calls", 0) > 0:
        row += 1
        ws_summary.cell(row=row, column=1, value="AI Token Usage").font = SUBHEADER_FONT
        row += 1

        token_stats = [
            ("API Calls", str(token_usage.get("api_calls", 0))),
            ("Input Tokens", f"{token_usage.get('prompt_tokens', 0):,}"),
            ("Output Tokens", f"{token_usage.get('completion_tokens', 0):,}"),
            ("Thinking Tokens", f"{token_usage.get('thinking_tokens', 0):,}"),
            ("Total Tokens", f"{token_usage.get('total_tokens', 0):,}"),
        ]

        for label, value in token_stats:
            ws_summary.cell(row=row, column=1, value=label).font = BODY_FONT
            ws_summary.cell(row=row, column=2, value=value).font = Font(name="Roboto", bold=True, size=10)
            row += 1

    # Column widths for summary
    ws_summary.column_dimensions["A"].width = 50
    ws_summary.column_dimensions["B"].width = 30
    ws_summary.column_dimensions["C"].width = 18
    ws_summary.column_dimensions["D"].width = 18
    ws_summary.column_dimensions["E"].width = 18
    ws_summary.column_dimensions["F"].width = 14

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
